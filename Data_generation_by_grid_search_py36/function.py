import numpy as np

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def center(q, s):
    q_mean = np.dot(s, q)
    q_centered = q - q_mean
    return q_mean, q_centered


def get_mean_sd_var(a, s):
    a_mean, a_centered = center(a, s)
    # print(f"a_mean: {a_mean}")
    var_1 = np.multiply(s, a_centered)
    var = np.dot(a_centered, var_1)
    sd = np.sqrt(var)
    return a_mean, sd, var


def get_covar(a,b, s):
    q, a_centered = center(a, s)
    q, b_centered = center(b, s)
    covar_1 = np.multiply(s, a_centered)
    covar = np.dot(b_centered, covar_1)
    return covar


def get_mean_sd_var_skew_skewBL(a, s):
    a_mean, a_centered = center(a, s)
    var_1 = np.multiply(s, a_centered)
    var = np.dot(a_centered, var_1)
    sd = np.sqrt(var)
    skew_1 = np.multiply(var_1, a_centered)
    skew_BL = np.dot(skew_1, a_centered)
    skew=skew_BL/np.power(sd,3)
    return a_mean, sd, var, skew, skew_BL


def get_eu(profit_mean, profit_sd, profit_var, RA):
    return profit_mean - RA   *  profit_var



def get_total_eu(forward_position, producer, retailer, outcome):
    return get_eu(
        *get_mean_sd_var(
            producer.get_profits_expost(
                outcome, forward_position
            ),
            outcome.ss_corrected ),
        producer.RA

    ) + get_eu(
        *get_mean_sd_var(
            retailer.get_profits_expost(
                outcome, forward_position
            ),
            outcome.ss_corrected ),
        retailer.RA)